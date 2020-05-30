import openpyxl

class HomepageTestData:

    data = {"firstname":"shivani","email":"abc@abc.com","password":"abc","gender":"Female"},{"firstname":"ranu","email":"abc@abc.com","password":"abc","gender":"Female"}

    @staticmethod
    def getData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("D:\\Learning\\AutomationWIthPython\\exceldata.xlsx")
        sheet = book.active
        rowcount = sheet.max_row
        colcount = sheet.max_column
        for i in range(1, rowcount + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, colcount + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]

