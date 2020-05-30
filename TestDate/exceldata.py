import openpyxl

book=openpyxl.load_workbook("D:\\Learning\\AutomationWIthPython\\exceldata.xlsx")
sheet=book.active
data=sheet.cell(row=1,column=2)
print(data.value)
rowcount=sheet.max_row
colcount=sheet.max_column
for i in range(1,rowcount+1):
    if sheet.cell(row=i,column=1).value=="tc02":
        for j in range (2,colcount+1):
            print(sheet.cell(row=i,column=j).value)